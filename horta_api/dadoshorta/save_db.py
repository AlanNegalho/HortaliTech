# from django.http import HttpResponse
# from openpyxl import Workbook
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from .models import DadhosHorta, UmidadeSolo

# class ExportXLSX(APIView):
#     def get(self, request, format=None):
#         # Obtenha os dados dos modelos DadhosHorta e UmidadeSolo
#         dados_horta = DadhosHorta.objects.all()
#         umidade_solo = UmidadeSolo.objects.all()

#         # Crie um novo arquivo XLSX
#         workbook = Workbook()
#         planilha = workbook.active

#         # Crie cabeçalhos para as colunas
#         planilha.cell(row=1, column=1, value='Temperatura')
#         planilha.cell(row=1, column=2, value='Umidade')
#         planilha.cell(row=1, column=3, value='Data')

#         # Preencha a planilha com os dados
#         for linha, dado in enumerate(dados_horta, start=2):
#             planilha.cell(row=linha, column=1, value=dado.temperatura)
#             planilha.cell(row=linha, column=2, value=dado.umidade)
#             planilha.cell(row=linha, column=3, value=dado.data.strftime('%d %m-%Y %H:%M:%S%'))  # Formate a data como string

#         for linha, dado in enumerate(umidade_solo, start=len(dados_horta) + 2):
#             planilha.cell(row=linha, column=1, value='')
#             planilha.cell(row=linha, column=2, value=dado.umidade)
#             planilha.cell(row=linha, column=3, value=dado.data.strftime('%d %m-%Y %H:%M:%S%'))  # Formate a data como string

#         # Configure a resposta para download do arquivo
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=dados_horta_umidade.xlsx'

#         # Salve o arquivo XLSX na resposta
#         workbook.save(response)

#         return response



# import zipfile
# from django.http import StreamingHttpResponse
# from wsgiref.util import FileWrapper
# from openpyxl import Workbook
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from .models import DadhosHorta, UmidadeSolo

# class ExportXLSX(APIView):
#    def get(self, request, format=None):
#        # Obtenha os dados dos modelos DadhosHorta e UmidadeSolo
#        dados_horta = DadhosHorta.objects.all()
#        umidade_solo = UmidadeSolo.objects.all()

#        # Crie dois arquivos XLSX separados
#        workbook_horta = Workbook()
#        planilha_horta = workbook_horta.active
#        workbook_solo = Workbook()
#        planilha_solo = workbook_solo.active

#        # Crie cabeçalhos para as colunas
#        planilha_horta.cell(row=1, column=1, value='Temperatura')
#        planilha_horta.cell(row=1, column=2, value='Umidade')
#        planilha_horta.cell(row=1, column=3, value='Data')

#        planilha_solo.cell(row=1, column=1, value='Umidade')
#        planilha_solo.cell(row=1, column=2, value='Data')

#        # Preencha os arquivos XLSX com os dados dos modelos
#        for linha, dado in enumerate(dados_horta, start=2):
#            planilha_horta.cell(row=linha, column=1, value=dado.temperatura)
#            planilha_horta.cell(row=linha, column=2, value=dado.umidade)
#            planilha_horta.cell(row=linha, column=3, value=dado.data.strftime('%d %m-%Y %H:%M:%S'))# Formate a data como string

#        for linha, dado in enumerate(umidade_solo, start=2):
#            planilha_solo.cell(row=linha, column=1, value=dado.umidade)
#            planilha_solo.cell(row=linha, column=2, value=dado.data.strftime('%d %m-%Y %H:%M:%S'))
#  # Formate a data como string

#        # Salve os arquivos XLSX
#        workbook_horta.save('dados_horta.xlsx')
#        workbook_solo.save('umidade_solo.xlsx')

#        # Crie um arquivo ZIP e adicione os arquivos XLSX a ele
#        with zipfile.ZipFile('dados_horta_umidade.zip', 'w') as zip_file:
#            zip_file.write('dados_horta.xlsx')
#            zip_file.write('umidade_solo.xlsx')

#        # Configure a resposta para download do arquivo ZIP
#        response = StreamingHttpResponse(FileWrapper(open('dados_horta_umidade.zip', 'rb')), content_type='application/zip')
#        response['Content-Disposition'] = 'attachment; filename=dados_horta_umidade.zip'

#        return response


import zipfile
from django.http import StreamingHttpResponse
from wsgiref.util import FileWrapper
from openpyxl import Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import DadhosHorta, UmidadeSolo

class ExportXLSX(APIView):
  def get(self, request, format=None):
      # Obtenha os dados dos modelos DadhosHorta e UmidadeSolo
      dados_horta = DadhosHorta.objects.all()
      umidade_solo = UmidadeSolo.objects.all()

      # Crie dois arquivos XLSX separados
      workbook_horta = Workbook()
      planilha_horta = workbook_horta.active
      workbook_solo = Workbook()
      planilha_solo = workbook_solo.active

      # Crie cabeçalhos para as colunas
      planilha_horta.cell(row=1, column=1, value='Temperatura')
      planilha_horta.cell(row=1, column=2, value='Umidade')
      planilha_horta.cell(row=1, column=3, value='Data')

      planilha_solo.cell(row=1, column=1, value='Umidade')
      planilha_solo.cell(row=1, column=2, value='Data')

      # Preencha os arquivos XLSX com os dados dos modelos
      for linha, dado in enumerate(dados_horta, start=2):
          if dado.temperatura is None or dado.umidade is None or dado.data is None:
              continue
          planilha_horta.cell(row=linha, column=1, value=dado.temperatura)
          planilha_horta.cell(row=linha, column=2, value=dado.umidade)
          planilha_horta.cell(row=linha, column=3, value=dado.data.strftime('%d %m-%Y %H:%M:%S'))

      for linha, dado in enumerate(umidade_solo, start=len(dados_horta) + 2):
          if dado.umidade is None or dado.data is None:
              continue
          planilha_solo.cell(row=linha, column=1, value=dado.umidade)
          planilha_solo.cell(row=linha, column=2, value=dado.data.strftime('%d %m-%Y %H:%M:%S'))

      # Salve os arquivos XLSX
      workbook_horta.save('dados_horta.xlsx')
      workbook_solo.save('umidade_solo.xlsx')

      # Crie um arquivo ZIP e adicione os arquivos XLSX a ele
      with zipfile.ZipFile('dados_horta_umidade.zip', 'w') as zip_file:
          zip_file.write('dados_horta.xlsx')
          zip_file.write('umidade_solo.xlsx')

      # Configure a resposta para download do arquivo ZIP
      response = StreamingHttpResponse(FileWrapper(open('dados_horta_umidade.zip', 'rb')), content_type='application/zip')
      response['Content-Disposition'] = 'attachment; filename=dados_horta_umidade.zip'

      return response
